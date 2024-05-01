"""This module contains various functions related to generating the external files."""

# pylint: disable=protected-access
# Standard library imports
import ast

# Public library imports
from openpyxl.utils import get_column_letter
import pandas as pd

# Personal library imports
# Module imports
# Initialized variables


def create_csv(input_data, filename):
    """Creates a CSV file."""
    input_data.to_csv(filename, index=False)


def create_xlsx(input_data, filename):
    """Creates an Excel file.
    Auto sizes columns and rows.
    Word wraps."""
    excel_writer = pd.ExcelWriter(filename)
    for item in input_data:
        input_data[item] = input_data[item].reset_index(drop=True)
        input_data[item].to_excel(excel_writer, sheet_name=item, index=False)

    # Auto-size columns
    for sheet_name in excel_writer.sheets:
        worksheet = excel_writer.sheets[sheet_name]
        worksheet.auto_filter.ref = worksheet.dimensions
        for col_num, col in enumerate(worksheet.iter_cols()):
            max_col_width = 0
            for cell in col:
                max_col_width = max(max_col_width, len(str(cell.value)))
            max_col_width = min(max_col_width + 7, 30)  # Cap at 30 characters (+2 padding)
            worksheet.column_dimensions[get_column_letter(col_num + 1)].width = max_col_width
        # Apply text wrapping
        for row in worksheet.iter_rows(min_row=2):  # Start from the second row (header is row 1)
            for cell in row:
                new_alignment = cell.alignment.copy()  # Create a copy
                new_alignment.wrap_text = True  # Modify wrap_text on the copy
                cell.alignment = new_alignment

                # Newline insertion (conditional)
                if cell.value.startswith("[") and cell.value.endswith("]"):
                    list_version = ast.literal_eval(cell.value)  # Convert to list
                    cell.value = "\n".join(list_version)  # Convert back to string with newlines

    excel_writer._save()


def create_all_reports(extracted_data, config):
    """Creates all reports, in CSV, XLSX, or both types. XLSX will contain all data, where CSV will have a seperate file per data point.
    Requires passing of configuration for the export type as a dictionary item "export_type" = "CSV" or "XLSX" or "BOTH"."""
    if config["export_type"] in ["CSV", "BOTH"]:
        for item in extracted_data:
            if config.get("output_directory"):
                create_csv(extracted_data[item], f"{config['output_directory']}\\{item}_report.csv")
            else:
                create_csv(extracted_data[item], f"{item}_report.csv")

    if config["export_type"] in ["XLSX", "BOTH"]:
        if config.get("output_directory"):
            create_xlsx(extracted_data, f"{config['output_directory']}\\ad_report.xlsx")
        else:
            create_xlsx(extracted_data, "ad_report.xlsx")
