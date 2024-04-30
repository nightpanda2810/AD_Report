"""This module contains various functions related to extracting and/or formatting data from Active Directory, as well as generating the external files."""

# pylint: disable=protected-access
# Standard library imports
import ast
import re

# Public library imports
from openpyxl.utils import get_column_letter
import pandas as pd

# Personal library imports
# Module imports
# Initialized variables
CN_PATTERN = r"CN=([^,]+)"
FLAG_MAP = {
    "Account_Disabled": 0x2,
    "Pwd_Not_Required": 0x20,
    "Pwd_Cant_Change": 0x40,
    "Pwd_Doesnt_Expire": 0x10000,
}


def extract_ad_data(input_data, the_attributes):
    """Extracts all requested data from Active Directory."""
    data = []
    for entry in input_data:
        row = {}
        for attr in the_attributes:
            try:
                if attr == "userAccountControl":
                    row.update(extract_user_account_control(entry, FLAG_MAP))
                elif attr in ["pwdLastSet", "modifyTimeStamp", "accountExpires", "createTimeStamp", "lastLogon"]:
                    attr_name = attr if attr != "modifyTimeStamp" else "Date_Modified"
                    row[attr_name] = extract_timestamp(entry[attr].values[0])
                elif attr == "member":
                    row["member"] = extract_member(entry[attr].values)
                elif attr == "msDS-parentdistname":
                    row["Parent_Location"] = entry[attr].values[0]
                else:
                    row[attr] = entry[attr].values[0]
            except (AttributeError, IndexError):
                row[attr] = ""
        data.append(row)
    return data


def extract_user_account_control(entry, flags):
    """Helper function for user_account_control attribute. Splits the user account control flags into seperate columns and returns a dictionary."""
    col = {column_name: "N/A" for column_name in flags.keys()}
    if is_user_enabled(entry):
        col["Account_Disabled"] = "Yes"
    else:
        col["Account_Disabled"] = "No"
    if is_password_not_required(entry):
        col["Pwd_Not_Required"] = "Yes"
    else:
        col["Pwd_Not_Required"] = "No"
    if is_password_cant_change(entry):
        col["Pwd_Cant_Change"] = "Yes"
    else:
        col["Pwd_Cant_Change"] = "No"
    if is_password_doesnt_expire(entry):
        col["Pwd_Doesnt_Expire"] = "Yes"
    else:
        col["Pwd_Doesnt_Expire"] = "No"
    return col


def extract_timestamp(value):
    """Helper function to convert datetime attributes into a date. Also converts 1601-01-01 and 999-12-31 into "Never" when they show up."""
    if str(value)[:10] in ["1601-01-01", "9999-12-31"]:
        return "Never"
    return str(value)[:10]


def extract_member(value):
    """Helper function to convert member attribute into a list, or blank string if appropiate."""
    new_data = []
    for item in value:
        new_value = re.search(CN_PATTERN, item)
        if new_value:
            new_data.append(new_value.group(1))
    return new_data if new_data else ""


def extract_and_format_ad_data(input_data, attributes):
    """Extracts desired attributes and creates a DataFrame."""
    data = extract_ad_data(input_data, attributes)  # Reuse your existing function
    return pd.DataFrame(data)


def search_and_extract_data(base_dn, search_filter, attributes, the_connection):
    """Extracts required data."""
    the_connection.search(base_dn, search_filter, attributes=attributes)
    entries = the_connection.entries
    return extract_and_format_ad_data(entries, attributes)


def is_user_enabled(entry):
    """Checks if the account is enabled."""
    uac = entry["userAccountControl"][0]  # Get the userAccountControl value
    return uac & 0x2  #  Check if the flag is NOT set


def is_password_not_required(entry):
    """Checks if the account requires a password."""
    uac = entry["userAccountControl"][0]
    return uac & 0x20


def is_password_cant_change(entry):
    """Checks if the account cannot change its password."""
    uac = entry["userAccountControl"][0]
    return uac & 0x40


def is_password_doesnt_expire(entry):
    """Checks if the account password does not expire."""
    uac = entry["userAccountControl"][0]
    return uac & 0x10000


def create_csv(input_data, filename):
    """Creates a CSV file."""
    input_data.to_csv(filename, index=False)


def create_xlsx(input_data, filename):
    """Creates an Excel file."""
    excel_writer = pd.ExcelWriter(filename)
    for item in input_data:
        input_data[item] = input_data[item].reset_index(drop=True)
        input_data[item].to_excel(excel_writer, sheet_name=item, index=False)

    # Auto-size columns
    for sheet_name in excel_writer.sheets:
        worksheet = excel_writer.sheets[sheet_name]
        # Auto-size columns
        for col_num, col in enumerate(worksheet.iter_cols()):
            max_col_width = 0
            for cell in col:
                max_col_width = max(max_col_width, len(str(cell.value)))
            max_col_width = min(max_col_width + 5, 30)  # Cap at 30 characters (+2 padding)
            worksheet.column_dimensions[get_column_letter(col_num + 1)].width = max_col_width
        # Apply text wrapping
        for row in worksheet.iter_rows(min_row=2):  # Start from the second row (header is row 1)
            for cell in row:
                new_alignment = cell.alignment.copy()  # Create a copy
                new_alignment.wrap_text = True  # Modify wrap_text on the copy
                cell.alignment = new_alignment

                # Newline insertion (conditional)
                if cell.value.startswith("[") and cell.value.endswith("]"):
                    list_version = ast.literal_eval(cell.value)  # Convert to list
                    cell.value = "\n".join(list_version)  # Convert back to string with newlines

    excel_writer._save()


def create_all_reports(extracted_data, config):
    """Creates all reports, in CSV, XLSX, or both types. XLSX will contain all data, where CSV will have a seperate file per data point.
    Requires passing of configuration for the export type as a dictionary item "export_type" = "CSV" or "XLSX" or "BOTH"."""
    if config["export_type"] in ["CSV", "BOTH"]:
        for item in extracted_data:
            if config.get("output_directory"):
                create_csv(extracted_data[item], f"{config['output_directory']}\\{item}_report.csv")
            else:
                create_csv(extracted_data[item], f"{item}_report.csv")

    if config["export_type"] in ["XLSX", "BOTH"]:
        if config.get("output_directory"):
            create_xlsx(extracted_data, f"{config['output_directory']}\\ad_report.xlsx")
        else:
            create_xlsx(extracted_data, "ad_report.xlsx")
